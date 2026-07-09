"""
AI parsing engine — turns free-text shift notes or uploaded registry files
into structured commissioning records.

Uses Groq's hosted API rather than local Ollama, so this works when deployed
to Streamlit Community Cloud (there's no localhost Ollama server there).
Requires GROQ_API_KEY in Streamlit secrets.
"""
import io
import re
import json
import hashlib
import requests
import streamlit as st

from config import MILESTONES, SCOPE_MILESTONES
from database import upsert_registry_row, check_chunk_exists, mark_chunk_done, \
    upload_file_to_storage, record_file_metadata

CHUNK_CHAR_LIMIT = 30000

SYSTEM_PROMPT = f"""
You are a commissioning data extractor for a nuclear power plant Reactor Shop.

KKS TAXONOMY:
- System KKS: a 3-letter code (e.g. JEA, JAA, JEC).
- Equipment KKS: a 2-letter code (e.g. AA, AP).

COMMISSIONING MILESTONES: {", ".join(m.upper() for m in MILESTONES)}
- IT: Individual Test
- PIC: Post Installation Cleaning / Flushing
- HT: Hydro Test
- PT: Pneumatic Test
- SAW: Start-up and Adjustment Work

RULES:
1. If the KKS prefix has 3 letters, scope_type = "System" (all five milestones apply).
   If it has 2 letters, scope_type = "Equipment" (pt_status and saw_status must be "N/A").
2. PIC (flushing) should normally be completed before HT (hydro test) is meaningful —
   if a note reports HT progress while PIC is not Completed, still record what the
   note says, but do not invent a PIC status that wasn't mentioned.
3. Only set a milestone status field if the note actually addresses that milestone.
   Leave unmentioned fields out of the JSON entirely (do not guess).
4. Do not create duplicate records — every record is matched by (system, component).

LANGUAGE:
Source notes and spreadsheets from this plant are frequently in Russian, English,
or a mix of both in the same row (e.g. bilingual column headers, Cyrillic status
words like "Выполнено"). Read and understand Russian text. Always translate
`comments` into English, and always output status values using the English
canonical vocabulary below — never leave a Russian word in a status field.

Extract data to a single JSON object (or a JSON object with a "records" list for
multiple items) using only these keys: system, system_kks, scope_type, component,
it_status, pic_status, ht_status, pt_status, saw_status, comments.
Status values must be one of: Pending, In Progress, Completed, Failed, N/A.
Output JSON only — no markdown, no explanation.
"""


# Deterministic safety net: even though the prompt instructs the model to
# translate statuses into English, this catches any Russian term that slips
# through un-translated so it still lands in a valid canonical status rather
# than being silently rejected downstream.
RU_STATUS_MAP = {
    "выполнено": "Completed", "завершено": "Completed", "выполнен": "Completed",
    "в процессе": "In Progress", "выполняется": "In Progress", "в работе": "In Progress",
    "не начато": "Pending", "не выполнено": "Pending", "ожидание": "Pending",
    "отклонено": "Failed", "не пройдено": "Failed", "неудачно": "Failed",
    "не применимо": "N/A", "н/п": "N/A",
}


def normalize_status(value: str) -> str:
    """Maps a Russian status term to its English canonical equivalent, if
    recognized. Leaves already-English values (or unrecognized ones)
    untouched — validation against STATUS_OPTIONS happens by the caller."""
    if not value:
        return value
    key = str(value).strip().lower()
    return RU_STATUS_MAP.get(key, value)


def get_kks_scope(kks_code: str) -> str:
    """
    Determines System vs Equipment scope from KKS prefix length, per the
    ROLE spec: 3-letter prefix = System, 2-letter prefix = Equipment.

    (The previous implementation used a hardcoded whitelist of specific
    prefixes, which silently misclassified anything not on the list. This
    checks the actual letter-length rule instead.)
    """
    if not kks_code:
        return "Equipment"
    first_token = str(kks_code).split(",")[0].strip()
    letters = re.sub(r"[^A-Za-z]", "", first_token).upper()
    if len(letters) == 3:
        return "System"
    if len(letters) == 2:
        return "Equipment"
    # Ambiguous length (not 2 or 3 letters) — default to the safer,
    # more restrictive scope and let the reviewer correct it if wrong.
    return "Equipment"


def enforce_scope_rules(record: dict) -> list[str]:
    """
    Applies scope-based N/A enforcement and flags milestone-dependency
    issues (PIC should precede HT). Mutates `record` in place to correct
    invalid milestone/scope combinations, and returns a list of
    human-readable alerts to show the user.
    """
    alerts = []
    scope = record.get("scope_type") or get_kks_scope(record.get("system_kks", ""))
    record["scope_type"] = scope
    applicable = SCOPE_MILESTONES.get(scope, SCOPE_MILESTONES["Equipment"])

    # Catch any Russian status term the model didn't translate.
    for m in MILESTONES:
        key = f"{m}_status"
        if record.get(key):
            record[key] = normalize_status(record[key])

    for m in MILESTONES:
        key = f"{m}_status"
        if m not in applicable and record.get(key) not in (None, "", "N/A"):
            alerts.append(
                f"⚠️ '{m.upper()}' was set to '{record.get(key)}' on "
                f"'{record.get('component', '(unnamed)')}', but {m.upper()} is N/A "
                f"for Equipment-scope items. Overriding to N/A."
            )
            record[key] = "N/A"

    ht_val = record.get("ht_status")
    pic_val = record.get("pic_status")
    if ht_val in ("In Progress", "Completed") and pic_val not in ("Completed", None, ""):
        if pic_val != "Completed":
            alerts.append(
                f"⚠️ HT is being marked '{ht_val}' for '{record.get('component', '(unnamed)')}' "
                f"but PIC (flushing) is currently '{pic_val}', not Completed. "
                f"Verify flushing was actually finished before relying on this hydro test."
            )

    if not record.get("system") or not record.get("component"):
        alerts.append(
            "⚠️ A parsed record is missing 'system' or 'component' — it was skipped "
            "because upserts require both to identify the row."
        )
    return alerts


def ask_groq(prompt: str):
    """Groq API wrapper with JSON enforcement and proper error handling."""
    api_key = st.secrets.get("GROQ_API_KEY")
    if not api_key:
        st.error("GROQ_API_KEY is missing from Streamlit Secrets — the AI parser can't run without it.")
        return None

    payload = {
        "model": "llama-3.3-70b-versatile",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        "response_format": {"type": "json_object"},
    }
    try:
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            json=payload,
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=30,
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        return json.loads(content)
    except requests.exceptions.RequestException as exc:
        st.error(f"Groq API request failed: {exc}")
    except (KeyError, json.JSONDecodeError) as exc:
        st.error(f"Groq returned an unexpected response format: {exc}")
    return None


def _rows_as_text(file_bytes: bytes, file_name: str) -> str:
    """
    Converts an uploaded file into readable text for the LLM prompt.

    The previous version called file_bytes.decode('utf-8') on EVERY file,
    including .xlsx — but .xlsx is a binary zip archive, not text, so that
    produced garbage input for any Excel upload. This branches by file type
    and reads .xlsx properly via openpyxl.
    """
    if file_name.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if any(v not in (None, "") for v in row):
                    lines.append(" | ".join(str(v) for v in row if v is not None))
        return "\n".join(lines)
    return file_bytes.decode("utf-8", errors="ignore")


def _chunk_text(text: str, char_limit: int = CHUNK_CHAR_LIMIT) -> list[str]:
    """Chunks on line boundaries instead of raw character slicing, so a
    row/record is never split in half across two chunks."""
    chunks, current, current_len = [], [], 0
    for line in text.split("\n"):
        if current_len + len(line) > char_limit and current:
            chunks.append("\n".join(current))
            current, current_len = [], 0
        current.append(line)
        current_len += len(line) + 1
    if current:
        chunks.append("\n".join(current))
    return chunks or [text]


def parse_shift_note(note_text: str) -> dict:
    """
    Parses a single free-text shift/field note into one or more staged
    commissioning records. Unlike process_file_smart, this does NOT save
    anything — it returns records for the caller to show in a review/confirm
    UI first, since a shift note is a live human report and worth a quick
    eyeball before it overwrites registry data.

    Returns: {"records": list[dict], "alerts": list[str]}
    """
    if not note_text or not note_text.strip():
        return {"records": [], "alerts": ["Note is empty."]}

    data = ask_groq(note_text)
    if not data:
        return {"records": [], "alerts": ["The AI parser didn't return a usable response. Try rephrasing the note, or fill it in manually."]}

    records = data.get("records", [data]) if isinstance(data, dict) else data
    all_alerts = []
    for record in records:
        record["scope_type"] = record.get("scope_type") or get_kks_scope(record.get("system_kks", ""))
        all_alerts.extend(enforce_scope_rules(record))

    return {"records": records, "alerts": all_alerts}


def process_file_smart(file_bytes: bytes, file_name: str) -> dict:
    """
    Incremental processing engine: stores the raw file in Supabase Storage,
    splits its content into row-safe chunks, skips chunks already processed
    (by file hash + chunk index) to avoid re-billing the LLM on re-upload,
    and upserts every extracted record — applying scope/dependency rules
    to each one.

    Returns a summary dict: {"records_saved": int, "alerts": list[str], "chunks_skipped": int}
    """
    file_hash = hashlib.md5(file_bytes).hexdigest()
    storage_path = upload_file_to_storage(file_bytes, file_name)

    text = _rows_as_text(file_bytes, file_name)
    chunks = _chunk_text(text)

    records_saved = 0
    chunks_skipped = 0
    all_alerts = []

    for i, chunk in enumerate(chunks):
        if check_chunk_exists(file_hash, i):
            chunks_skipped += 1
            continue

        data = ask_groq(chunk)
        if not data:
            all_alerts.append(f"⚠️ Chunk {i + 1}/{len(chunks)} failed to parse and was skipped.")
            continue

        records = data.get("records", [data]) if isinstance(data, dict) else data

        for record in records:
            record["scope_type"] = record.get("scope_type") or get_kks_scope(record.get("system_kks", ""))
            alerts = enforce_scope_rules(record)
            all_alerts.extend(alerts)

            if not record.get("system") or not record.get("component"):
                continue  # already alerted above; can't upsert without the conflict key

            record["source"] = file_name
            upsert_registry_row(record)
            records_saved += 1

        mark_chunk_done(file_hash, i)

    if storage_path:
        record_file_metadata(file_name, storage_path, records_saved)

    return {"records_saved": records_saved, "alerts": all_alerts, "chunks_skipped": chunks_skipped}
